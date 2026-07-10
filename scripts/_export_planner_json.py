"""Extrae radicado + enlaces del export Planner → JSON stdout."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"c:\Users\cmarroqh\Downloads\Procesos Juzgado 51 Ccto.xlsx")
sheet = sys.argv[2] if len(sys.argv) > 2 else "Datos consolidados"
extra_args = sys.argv[4:] if len(sys.argv) > 4 else []
solo_activas = "--todas" not in extra_args and "--incluir-completadas" not in extra_args

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
ws = wb[sheet]

headers: dict[str, int] = {}
for col in range(1, ws.max_column + 1):
    v = ws.cell(1, col).value
    if isinstance(v, str) and v.strip():
        headers[v.strip()] = col

name_col = headers.get("Nombre de la tarea")
notes_col = headers.get("Notas")
bucket_col = headers.get("Depósito") or headers.get("Dep\u00f3sito")
state_col = headers.get("Estado")
tags_col = headers.get("Etiquetas")
due_col = headers.get("Fecha de vencimiento")

rad_re = re.compile(r"\d{21,23}")
share_re = re.compile(r"sharepoint|onedrive|1drv\.ms", re.I)
sgde_re = re.compile(r"add-ficheros/([0-9a-f-]{36})", re.I)


def normalize_radicado(raw: str) -> str:
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) >= 23:
        return digits[:23]
    if len(digits) >= 21:
        return digits.zfill(23)[:23]
    return ""


def pick_radicado(task_name: str, notes: str) -> str:
    candidates: list[str] = []
    if notes:
        first = notes.split("\n", 1)[0].strip()
        if first:
            candidates.append(first)
    if task_name:
        candidates.append(task_name)
    for src in candidates:
        m = rad_re.search(src or "")
        if m:
            r = normalize_radicado(m.group(0))
            if len(r) >= 21:
                return r
    return ""


def is_planner_estado_activo(estado: str) -> bool:
    e = (estado or "").strip().lower()
    if not e:
        return False
    if e in ("completado", "completed", "completo", "done"):
        return False
    return e in ("en curso", "no iniciado", "not started", "in progress")


def urls_from_cell(cell) -> list[str]:
    out: list[str] = []
    v = cell.value
    if isinstance(v, str):
        for m in re.findall(r"https?://\S+", v):
            out.append(m.rstrip(".,;)"))
    hl = getattr(cell, "hyperlink", None)
    if hl and getattr(hl, "target", None):
        out.append(str(hl.target).strip())
    seen: set[str] = set()
    deduped: list[str] = []
    for u in out:
        if u not in seen:
            seen.add(u)
            deduped.append(u)
    return deduped


rows = []
skipped_completadas = 0
for r in range(2, ws.max_row + 1):
    estado = str(ws.cell(r, state_col).value or "").strip() if state_col else ""
    if solo_activas and not is_planner_estado_activo(estado):
        skipped_completadas += 1
        continue
    task_name = str(ws.cell(r, name_col).value or "").strip() if name_col else ""
    notes_cell = ws.cell(r, notes_col) if notes_col else None
    notes = str(notes_cell.value or "").strip() if notes_cell else ""
    urls = urls_from_cell(notes_cell) if notes_cell else []
    share = next((u for u in urls if share_re.search(u)), None)
    sgde = next((u for u in urls if sgde_re.search(u)), None)
    sgde_id = None
    if sgde:
        m = sgde_re.search(sgde)
        if m:
            sgde_id = m.group(1).lower()
    rad = pick_radicado(task_name, notes)
    if not rad and not share and not sgde:
        continue
    rows.append(
        {
            "radicado": rad,
            "task_name": task_name,
            "link_expediente": share,
            "sgde_url": sgde,
            "sgde_id": sgde_id,
            "deposito": str(ws.cell(r, bucket_col).value or "").strip() if bucket_col else "",
            "estado_planner": estado,
            "etiquetas": str(ws.cell(r, tags_col).value or "").strip() if tags_col else "",
            "fecha_vencimiento": str(ws.cell(r, due_col).value or "").strip() if due_col else "",
            "notas_preview": notes[:500] if notes else "",
        }
    )

stats = {
    "solo_activas": solo_activas,
    "skipped_completadas": skipped_completadas,
    "total_rows": len(rows),
    "with_radicado": sum(1 for x in rows if x["radicado"]),
    "with_sharepoint": sum(1 for x in rows if x["link_expediente"]),
    "with_sgde": sum(1 for x in rows if x["sgde_id"]),
}
out_path = Path(sys.argv[3]) if len(sys.argv) > 3 else None
payload = {"stats": stats, "rows": rows}
if out_path:
    out_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
else:
    json.dump(payload, sys.stdout, ensure_ascii=False)
